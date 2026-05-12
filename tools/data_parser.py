import openpyxl
from pathlib import Path
import pandas as pd
from datetime import datetime
import warnings
import sys 

# Ignore warnings 
warnings.filterwarnings('ignore', category=UserWarning)

def get_cell_value(ws, row, col): 
    try: 
        value = ws.cell(row=row, column=col).value
        if isinstance(value, datetime):
           return value.strftime("%Y-%m-%d")
        return value
    except:
        return None
    

def extract_datoscotizacion(wb):
    # Initialization 
    data = {}

    # Data Validation
    if "DatosCotización" not in wb.sheetnames:
        return data
    
    ws = wb["DatosCotización"]

    # Cell map definitions

    data['archivo_nombre'] = get_cell_value(ws, 4, 3)                   # C4
    data['plaza'] = get_cell_value(ws, 6, 3)                            # C6
    data['region'] = get_cell_value(ws, 6, 5)                           # E6
    data['codigo_ubicacion'] = get_cell_value(ws, 7, 3)                 # C7
    data['codigo_region'] = get_cell_value(ws, 7, 5)                    # E7

    data['cliente'] = get_cell_value(ws, 10, 3)                         # C10
    data['rfc'] = get_cell_value(ws, 11, 3)                             # C11
    data['contacto'] = get_cell_value(ws, 12, 3)                        # C6

    data['numero_concurso'] = get_cell_value(ws, 13, 3)                 # C13
    data['fecha_entrega'] = get_cell_value(ws, 13, 5)                   # E13
    data['centro'] = get_cell_value(ws, 14, 3)                          # C14
    data['domicilio'] = get_cell_value(ws, 15, 3)                       # C15
    data['poblacion'] = get_cell_value(ws, 16, 3)                       # C16
    data['estado'] = get_cell_value(ws, 17, 3)                          # C17

    data['nombre_proyecto'] = get_cell_value(ws, 20, 3)                 # C20
    data['tipo_servicio'] = get_cell_value(ws, 21, 3)                   # C21

    data['condiciones_pago'] = get_cell_value(ws, 23, 3)                # C23
    data['mb_minimo'] = get_cell_value(ws, 23, 5)                       # E23
    data['fianza_cumplimiento'] = get_cell_value(ws, 24, 5)             # E24
    data['numero_oferta'] = get_cell_value(ws, 25, 3)                   # C25
    data['seguro_rc'] = get_cell_value(ws, 25, 5)                       # E25
    data['impuesto_nominas'] = get_cell_value(ws, 26, 5)                # E26
    data['duracion_meses'] = get_cell_value(ws, 33, 3)                  # C33

    return data 

def extract_resumencotizacion(wb): 
    #Initialization
    data = {}

    # Data Validation
    if "ResumenCotización" not in wb.sheetnames: 
        return data
    
    ws = wb["ResumenCotización"]
    
    data['resumen_cliente'] = get_cell_value(ws, 4, 5)                  # E4
    data['resumen_centro'] = get_cell_value(ws, 5, 5)                   # E5
    data['resumen_servicio'] = get_cell_value(ws, 6, 3)                 # E9

    empleados_data = {}
    total_empleados = 0

    for row in range(15, 27):
        num_empleados = get_cell_value(ws, row, 3)
        categoria = get_cell_value(ws, row, 4)
        importe_anual = get_cell_value(ws, row, 5)
        importe_mensual = get_cell_value(ws, row, 6)

        if num_empleados and num_empleados != 0 and categoria: 
            empleados_data.append({
                'categoria': categoria, 
                'num_empleados': num_empleados, 
                'importe_anual': importe_anual, 
                'importe_mensual': importe_mensual
            })
            if isinstance(num_empleados, (int, float)):
                total_empleados += int( num_empleados)
    
    data['empleados_detalle'] = empleados_data
    data['total_empleados'] = total_empleados

    data['total_mano_obra_anual'] = get_cell_value(ws, 27, 5)           # E27
    data['total_mano_obra_anual'] = get_cell_value(ws, 27, 6)           # F27
    data['vehiculos_anual'] = get_cell_value(ws, 28, 5)                 # E28
    data['total_costo_servicio'] = get_cell_value(ws, 35, 5)            # E35

    return data 


def extract_edo_resultados(wb): 

    # Initialization
    data = {}

    # Data Validation
    if 'EdoResultados' not in wb.sheetnames: 
        return data
    
    ws = wb['EdoResultados']

    data['facturacion_anual'] = get_cell_value(ws, 17, 4)                # D17
    data['suma_ingresos_anual'] = get_cell_value(ws, 19, 4)              # D19

    data['costo_mano_obra_anual'] = get_cell_value(ws, 24, 4)            # D24


    return data


def extract_estudio_economico(wb): 
    # Initialization
    data = {}

    # Data Validation
    if 'EstudioEconomico' not in wb.sheetnames: 
        return data 
    
    ws = wb['EstudioEconomico']

    data['ee_proyecto'] = get_cell_value(ws, 13, 3)                     # C13
    data['ee_plaza'] = get_cell_value(ws, 14, 3)                        # C14
    data['ee_region'] = get_cell_value(ws, 14, 5)                       # E14
    data['ee_cliente'] = get_cell_value(ws, 16, 3)                      # C16
    data['ee_centro'] = get_cell_value(ws, 17, 3)                       # C17
    data['ee_duracion'] = get_cell_value(ws, 17, 7)                     # G17
    data['ee_contrato'] = get_cell_value(ws, 19, 3)                     # C19
    data['ee_tipo_servicio'] = get_cell_value(ws, 21, 3)                # C21

    return data


def process_file(filepath): 

    try: 
        wb = openpyxl.load_workbook(filepath, data_only=True)

        # Extract data from each sheet 
        datos = extract_datoscotizacion(wb)
        resumen = extract_resumencotizacion(wb)
        edo = extract_edo_resultados(wb)
        estudio = extract_estudio_economico(wb)

        wb.close()

        # Data combination
        combined = {
            'archivo': filepath.name,
            **datos,
            **{k: v for k, v in resumen.items() if k != 'empleados_detalle'},
            **edo,
            **estudio
        }

        empleados = resumen.get('empleados_detalle', [])

        return combined, empleados
    
    except Exception as e: 
        print(f"Error processing {filepath.name}: {e}")
        return None, []


def main(): 
    
    data_dir = Path('Datos/Datos')
    output_dir = Path(".")

    # Get excel files 
    xls_files = sorted(data_dir.glob('*.xlsm'))
    print(f"Found {len(xls_files)} files to process")

    all_records = []
    all_empleados = []
    errors = []

    for i, filepath in enumerate(xls_files, 1):

        record, empleados = process_file(filepath)

        if record:
            all_records.append(record)

            for emp in empleados: 
                emp['archivo'] = filepath.name
                emp['cliente'] = record.get('cliente', '')
                emp['centro'] = record.get('centro', '')
                emp['tipo_servicio'] = record.get('tipo_servicio', '')
                all_empleados.append(emp)
        else:
            errors.append(filepath.name)

    # Main dataframe
    df_main = pd.DataFrame(all_records)

    # Employees dataframe
    df_empleados = pd.DataFrame(all_empleados)

    # Save datset
    output_file = output_dir / 'Estudios_Economicos_Consolidado.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer: 
        df_main.to_excel(writer, sheet_name='Resumen_General', index=False)
        df_empleados.to_excel(writer, sheet_name='Detalle_empleados', index=False)
    
    # As CSV 
    df_main.to_csv(output_dir / 'Estudios_Economicos_Resumen.xlsx', index=False, encoding='utf-8-sig')
    df_empleados.to_csv(output_dir / 'Estudios_Economicos_Empleados.xlsx', index=False, encoding='utf-8-sig')

    print("Processing complete")


if __name__ == '__main__':
    main()